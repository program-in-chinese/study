from openpyxl import load_workbook

开始行 = 1
结束行 = 250

def 遍历表格(文件名):
    工作簿 = load_workbook(文件名)
    表单 = 工作簿.active
    for 行 in range(开始行, 结束行):
        代码 = 表单.cell(row = 行, column = 3).value
        短名称 = 表单.cell(row = 行, column = 1).value

        #名称长 = len(短名称)
        #if 名称长 > 5 and 短名称[短名称.length - 5 : 短名称.length]
        
        # 下面是格式化输出
        print("    '" + 代码 + "' : '" + 短名称 + "',")

遍历表格("源数据.xlsx")