from openpyxl import load_workbook

开始行 = 3
结束行 = 205

所有拼音 = []

def 遍历表格(文件名):
    工作簿 = load_workbook(文件名)
    表单 = 工作簿.active
    for 行 in range(开始行, 结束行):
        拼音1 = 表单.cell(row = 行, column = 1).value.lower()

        拼音1入声 = 表单.cell(row = 行, column = 2).value
        拼音1一声 = 表单.cell(row = 行, column = 3).value
        拼音1二声 = 表单.cell(row = 行, column = 4).value
        拼音1三声 = 表单.cell(row = 行, column = 5).value
        拼音1四声 = 表单.cell(row = 行, column = 6).value

        if 拼音1入声 != None:
            所有拼音.append(拼音1 + '0')
        if 拼音1一声 != None:
            所有拼音.append(拼音1 + '1')
        if 拼音1二声 != None:
            所有拼音.append(拼音1 + '2')
        if 拼音1三声 != None:
            所有拼音.append(拼音1 + '3')
        if 拼音1四声 != None:
            所有拼音.append(拼音1 + '4')

        拼音2 = 表单.cell(row = 行, column = 7).value.lower()

        拼音2入声 = 表单.cell(row = 行, column = 8).value
        拼音2一声 = 表单.cell(row = 行, column = 9).value
        拼音2二声 = 表单.cell(row = 行, column = 10).value
        拼音2三声 = 表单.cell(row = 行, column = 11).value
        拼音2四声 = 表单.cell(row = 行, column = 12).value
        
        if 拼音2入声 != None:
            所有拼音.append(拼音2 + '0')
        if 拼音2一声 != None:
            所有拼音.append(拼音2 + '1')
        if 拼音2二声 != None:
            所有拼音.append(拼音2 + '2')
        if 拼音2三声 != None:
            所有拼音.append(拼音2 + '3')
        if 拼音2四声 != None:
            所有拼音.append(拼音2 + '4')


遍历表格("汉语拼音四声大全.xlsx")
所有拼音.sort()
print(len(所有拼音))
for 拼音 in 所有拼音:
    print(拼音)